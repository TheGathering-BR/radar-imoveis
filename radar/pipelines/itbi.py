"""Pipeline do ITBI: download -> normalização -> gravação em `transacoes`.

Idempotente por aba: o conteúdo de cada aba mensal é hasheado; se não mudou
desde a última carga, a aba é pulada. Se mudou (caso do arquivo do ano
vigente, reescrito mensalmente), os registros antigos daquela aba são
substituídos.
"""
import hashlib
import itertools
import re
import sqlite3
import unicodedata
from collections import Counter
from datetime import date, datetime

import openpyxl
import requests

from radar.config import (
    ANO_VIGENTE,
    CIDADE_ATIVA,
    HTTP_HEADERS,
    ITBI_URLS,
    PRECO_M2_MAX,
    PRECO_M2_MIN,
    RAW_DIR,
)
from radar.geo import quadra_lookup

MESES = {"JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
         "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12}
RE_ABA_MES = re.compile(r"^([A-Z]{3})[-_ ]?(\d{4})$")


def _slug(texto: str) -> str:
    """Normaliza nome de coluna: sem acento, minúsculo, só [a-z0-9]."""
    s = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())


# Mapeia slug do cabeçalho -> campo interno (None = coluna reconhecida mas
# ignorada, para não colidir por prefixo). Cobre variações entre anos.
COLUNAS = {
    "ndocadastrosql": "sql",
    "numerodocadastro": "sql",
    "nomedologradouro": "logradouro",
    "numero": "numero",
    "complemento": "complemento",
    "bairro": "bairro_iptu",
    "cep": "cep",
    "naturezadetransacao": "natureza",
    "valordetransacao": "valor",
    "datadetransacao": "data",
    "valorvenaldereferenciaproporcional": None,
    "valorvenaldereferencia": "venal",
    "proporcaotransmitida": "proporcao",
    "basedecalculo": "base_calculo",
    "valorfinanciado": "financiado",
    "areadoterreno": "area_terreno",
    "areaconstruida": "area_construida",
    "fracaoideal": "fracao",
    "usoiptu": "uso",
    "descricaodouso": "desc_uso",
    "padraoiptu": "padrao",
    "descricaodopadrao": "desc_padrao",
    "acciptu": "ano_construcao",
}
# Chaves mais longas primeiro, para a mais específica vencer a genérica
# (ex.: "numerodocadastro" antes de "numero").
_COLUNAS_ORDENADAS = sorted(COLUNAS.items(), key=lambda kv: -len(kv[0]))


# Ordem canônica das 28 colunas, para abas que vêm sem linha de cabeçalho
# (aconteceu em JAN-2024 e OUT-2024). None = coluna ignorada.
ORDEM_CANONICA = [
    "sql", "logradouro", "numero", "complemento", "bairro_iptu", None,  # referência
    "cep", "natureza", "valor", "data", "venal", "proporcao",
    None,             # valor venal de referência (proporcional)
    "base_calculo",
    None,             # tipo de financiamento
    "financiado",
    None, None, None,  # cartório, matrícula, situação do SQL
    "area_terreno",
    None,             # testada
    "fracao", "area_construida",
    "uso", "desc_uso", "padrao", "desc_padrao", "ano_construcao",
]


def _mapear_colunas(header: tuple) -> dict:
    """Índice da coluna -> campo interno, tolerante a variações de grafia."""
    mapa = {}
    for i, nome in enumerate(header):
        if nome is None:
            continue
        slug = _slug(nome)
        for chave, campo in _COLUNAS_ORDENADAS:
            if slug.startswith(chave):
                if campo is not None:
                    mapa[i] = campo
                break
    return mapa


def _parece_linha_de_dados(row: tuple) -> bool:
    """Detecta aba sem cabeçalho: 1ª célula é um SQL numérico e a 8ª tem
    cara de natureza de transação ('1.Compra e venda' etc.)."""
    if not row or len(row) < len(ORDEM_CANONICA):
        return False
    if not isinstance(row[0], (int, float)):
        return False
    return isinstance(row[7], str) and bool(re.match(r"^\d+\.", row[7].strip()))


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if "," in s:  # formato brasileiro: '.' milhar, ',' decimal
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _data_iso(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def _texto(v):
    if v is None:
        return None
    # Células numéricas podem vir como float (ex.: nº da porta 295.0 em 2024)
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return s or None


def _digitos(v):
    """Extrai os dígitos de um valor que pode vir como int, float ou texto.

    Crítico para o SQL e o CEP: str(100800297.0) contém '.0' e um strip
    ingênuo de não-dígitos colaria um zero espúrio no final.
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if isinstance(v, int):
        return str(v)
    return re.sub(r"\D", "", str(v))


def baixar_ano(ano: int, force: bool = False):
    """Baixa o xlsx do ano para data/raw. O ano vigente é sempre re-baixado."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    destino = RAW_DIR / f"itbi_{ano}.xlsx"
    if destino.exists() and not force and ano != ANO_VIGENTE:
        return destino
    url = ITBI_URLS[ano]
    print(f"  baixando {url}")
    with requests.get(url, headers=HTTP_HEADERS, stream=True, timeout=600) as r:
        r.raise_for_status()
        tmp = destino.with_suffix(".part")
        with open(tmp, "wb") as f:
            for chunk in r.iter_content(chunk_size=1 << 20):
                f.write(chunk)
        tmp.replace(destino)
    return destino


def classificar_uso(descricao_uso) -> str | None:
    """Classe do imóvel a partir do uso IPTU.

    O IPTU só permite separar vertical (apartamento) de horizontal (casa);
    coberturas e casas de vila não são distinguíveis no ITBI. Usos não
    residenciais (garagem, loja, escritório...) retornam None e ficam fora
    de qualquer mediana.
    """
    if not descricao_uso:
        return None
    u = descricao_uso.upper()
    if u.startswith("APARTAMENTO"):
        return "apartamento"
    if u.startswith("RESIDÊNCIA COLETIVA") or u.startswith("RESIDENCIA COLETIVA"):
        return None  # mais de uma residência no lote — preço/m² não comparável
    if u.startswith("RESIDÊNCIA") or u.startswith("RESIDENCIA"):
        return "casa"  # inclui "RESIDÊNCIA E OUTRO USO (PREDOMINÂNCIA RESIDENCIAL)"
    return None


def _normalizar_registro(bruto: dict, mes_pagamento: str, quadras: dict) -> dict:
    sql_pad = None
    setor = quadra = None
    digitos = _digitos(bruto.get("sql"))
    if digitos and len(digitos) <= 11:
        sql_pad = digitos.zfill(11)
        setor, quadra = sql_pad[0:3], sql_pad[3:6]

    valor = _num(bruto.get("valor"))
    area_construida = _num(bruto.get("area_construida"))
    proporcao = _num(bruto.get("proporcao"))
    natureza = _texto(bruto.get("natureza"))

    lat = lon = bairro_id = None
    metodo = None
    if setor and quadra:
        hit = quadras.get((setor, quadra))
        if hit:
            lat, lon, bairro_id = hit
            metodo = "quadra_fiscal"

    # Filtros de mercado: compra e venda plena, com área construída válida.
    preco_m2 = None
    elegivel = 0
    eh_compra_venda = bool(natureza) and _slug(natureza).startswith("1compraevenda")
    if (eh_compra_venda and valor and valor > 0
            and area_construida and area_construida > 0
            and proporcao is not None and abs(proporcao - 100.0) < 0.01):
        preco_m2 = valor / area_construida
        if PRECO_M2_MIN <= preco_m2 <= PRECO_M2_MAX:
            elegivel = 1
        else:
            preco_m2 = None

    cep_dig = _digitos(bruto.get("cep"))
    cep = cep_dig.zfill(8) if cep_dig and len(cep_dig) <= 8 else None

    ano_c = _num(bruto.get("ano_construcao"))

    return {
        "cidade": CIDADE_ATIVA,
        "sql_cadastro": sql_pad,
        "logradouro": _texto(bruto.get("logradouro")),
        "numero": _texto(bruto.get("numero")),
        "complemento": _texto(bruto.get("complemento")),
        "cep": cep,
        "bairro_iptu": _texto(bruto.get("bairro_iptu")),
        "natureza": natureza,
        "valor_transacao": valor,
        "data_transacao": _data_iso(bruto.get("data")),
        "mes_pagamento": mes_pagamento,
        "valor_venal_ref": _num(bruto.get("venal")),
        "proporcao_pct": proporcao,
        "base_calculo": _num(bruto.get("base_calculo")),
        "valor_financiado": _num(bruto.get("financiado")),
        "area_terreno_m2": _num(bruto.get("area_terreno")),
        "area_construida_m2": area_construida,
        "fracao_ideal": _num(bruto.get("fracao")),
        "uso_iptu": _texto(bruto.get("uso")),
        "descricao_uso": _texto(bruto.get("desc_uso")),
        "padrao_iptu": _texto(bruto.get("padrao")),
        "descricao_padrao": _texto(bruto.get("desc_padrao")),
        "ano_construcao": int(ano_c) if ano_c else None,
        "lat": lat,
        "lon": lon,
        "geocode_metodo": metodo,
        "bairro_id": bairro_id,
        "preco_m2": preco_m2,
        "classe": classificar_uso(_texto(bruto.get("desc_uso"))),
        "elegivel_mediana": elegivel,
    }


CAMPOS_INSERT = [
    "cidade", "sql_cadastro", "logradouro", "numero", "complemento", "cep",
    "bairro_iptu", "natureza", "valor_transacao", "data_transacao",
    "mes_pagamento", "valor_venal_ref", "proporcao_pct", "base_calculo",
    "valor_financiado", "area_terreno_m2", "area_construida_m2",
    "fracao_ideal", "uso_iptu", "descricao_uso", "padrao_iptu",
    "descricao_padrao", "ano_construcao", "lat", "lon", "geocode_metodo",
    "bairro_id", "preco_m2", "classe", "elegivel_mediana",
    "fonte_arquivo", "fonte_aba", "hash_registro",
]


def ingerir_ano(conn: sqlite3.Connection, ano: int, quadras: dict,
                force_download: bool = False) -> dict:
    """Processa todas as abas mensais do arquivo de um ano."""
    arquivo = baixar_ano(ano, force=force_download)
    wb = openpyxl.load_workbook(arquivo, read_only=True)
    stats = {"abas_puladas": 0, "abas_carregadas": 0, "registros": 0}

    for nome_aba in wb.sheetnames:
        m = RE_ABA_MES.match(nome_aba.strip().upper())
        if not m or m.group(1) not in MESES:
            continue  # LEGENDA, EXPLICAÇÕES, tabelas auxiliares
        mes_pagamento = f"{m.group(2)}-{MESES[m.group(1)]:02d}"

        ws = wb[nome_aba]
        linhas = ws.iter_rows(values_only=True)
        header = next(linhas, None)
        if not header:
            continue
        mapa = _mapear_colunas(header)
        if "sql" not in mapa.values() or "valor" not in mapa.values():
            if _parece_linha_de_dados(header):
                # Aba sem cabeçalho: usa a ordem canônica e a 1ª linha é dado.
                mapa = {i: c for i, c in enumerate(ORDEM_CANONICA) if c}
                linhas = itertools.chain([header], linhas)
                print(f"  aba {nome_aba}: sem cabecalho, usando ordem canonica")
            else:
                print(f"  ! aba {nome_aba}: cabecalho nao reconhecido, pulando")
                continue

        brutos = []
        h = hashlib.sha1()
        for row in linhas:
            if row is None or all(v is None for v in row):
                continue
            h.update(repr(row).encode())
            brutos.append({campo: row[i] for i, campo in mapa.items() if i < len(row)})
        hash_aba = h.hexdigest()

        ja = conn.execute(
            "SELECT hash_conteudo FROM ingestoes WHERE fonte='itbi' AND arquivo=? AND aba=?",
            (arquivo.name, nome_aba),
        ).fetchone()
        if ja and ja[0] == hash_aba:
            stats["abas_puladas"] += 1
            continue

        # Substitui os registros da aba (arquivo do ano vigente muda todo mês).
        conn.execute(
            "DELETE FROM transacoes WHERE fonte_arquivo=? AND fonte_aba=?",
            (arquivo.name, nome_aba),
        )
        ocorrencias = Counter()
        registros = []
        for bruto in brutos:
            reg = _normalizar_registro(bruto, mes_pagamento, quadras)
            chave = (reg["sql_cadastro"], reg["data_transacao"],
                     reg["valor_transacao"], reg["natureza"])
            ocorrencias[chave] += 1
            reg["fonte_arquivo"] = arquivo.name
            reg["fonte_aba"] = nome_aba
            reg["hash_registro"] = hashlib.sha1(
                repr((chave, mes_pagamento, ocorrencias[chave])).encode()
            ).hexdigest()
            registros.append(tuple(reg[c] for c in CAMPOS_INSERT))

        placeholders = ",".join("?" * len(CAMPOS_INSERT))
        conn.executemany(
            f"INSERT OR IGNORE INTO transacoes ({','.join(CAMPOS_INSERT)}) "
            f"VALUES ({placeholders})",
            registros,
        )
        conn.execute(
            """INSERT INTO ingestoes (fonte, arquivo, aba, hash_conteudo, n_registros, processado_em)
               VALUES ('itbi', ?, ?, ?, ?, datetime('now'))
               ON CONFLICT (fonte, arquivo, aba) DO UPDATE SET
                 hash_conteudo=excluded.hash_conteudo,
                 n_registros=excluded.n_registros,
                 processado_em=excluded.processado_em""",
            (arquivo.name, nome_aba, hash_aba, len(registros)),
        )
        conn.commit()
        stats["abas_carregadas"] += 1
        stats["registros"] += len(registros)
        print(f"  aba {nome_aba}: {len(registros)} registros")

    wb.close()
    return stats


def ingerir(conn: sqlite3.Connection, anos, force_download: bool = False):
    quadras = quadra_lookup(conn)
    if not quadras:
        raise RuntimeError(
            "Tabela de quadras vazia — rode scripts/setup_db.py primeiro."
        )
    for ano in anos:
        print(f"[itbi] ano {ano}")
        stats = ingerir_ano(conn, ano, quadras, force_download=force_download)
        print(f"  -> {stats['abas_carregadas']} abas carregadas, "
              f"{stats['abas_puladas']} inalteradas, {stats['registros']} registros")
